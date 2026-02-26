from ungroup_util import ungroup_shapes_in_ppt
from flask import Flask, render_template, request, send_file
import os
import pandas as pd
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from animation_checker import run_animation_qc
from chunking_by_animation_win32 import run_chunking_qc_with_animation
from notes_validator import run_notes_validation
from text_rules_validator import run_text_rules_validation
from qc_points_generator import generate_qc_summary  # New import

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def clean_illegal_excel_chars(df):
    def clean_text(value):
        if isinstance(value, str):
            return ''.join([c for c in value if 32 <= ord(c) <= 126 or ord(c) in (9, 10, 13)])
        return value
    return df.applymap(clean_text)

def update_font_validation_with_fallback(excel_path):
    DEFAULT_STYLE_MAP = {
        ("Text Placeholder 2", "PLACEHOLDER (14)"): ("Queens Medium", 35),
        ("Text Placeholder 3", "PLACEHOLDER (14)"): ("HelveticaNowDisplay Medium", 27),
        ("Text Placeholder 14", "PLACEHOLDER (14)"): ("HelveticaNowDisplay Medium", 27),
    }

    allowed_fonts = {
        "HelveticaNowDisplay Medium": (24, 27),
        "Queens Medium": (35, 35),
        "HelveticaNowDisplay Black": (75, 92.5),
        "Consolas": (24, 35),
        "Cambria math": (24, 35),
    }

    allowed_font_colors = {"#000000", "#FFFFFF", "#F26722"}

    wb = load_workbook(excel_path)
    if "Quality Check" not in wb.sheetnames:
        return "Quality Check sheet not found."

    ws = wb["Quality Check"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx for idx, name in enumerate(header)}

    required_cols = ["Font Name", "Font Size", "Shape Name / Table Cell", "Shape Type", "Extracted Text", "Font Color Hex"]
    if not all(col in col_map for col in required_cols):
        return "Missing required columns."

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        font_cell = row[col_map["Font Name"]]
        size_cell = row[col_map["Font Size"]]
        shape_name = str(row[col_map["Shape Name / Table Cell"]].value or "").strip()
        shape_type = str(row[col_map["Shape Type"]].value or "").strip()
        extracted_text = str(row[col_map["Extracted Text"]].value or "").strip()
        font_color_cell = row[col_map["Font Color Hex"]]

        if not extracted_text:
            continue

        font = str(font_cell.value or "").strip()
        size_val = size_cell.value

        if not font:
            key = (shape_name, shape_type)
            fallback = DEFAULT_STYLE_MAP.get(key)
            if not fallback and shape_type == "TEXT_BOX (17)" and shape_name.lower().startswith("textbox"):
                fallback = ("HelveticaNowDisplay Medium", 27)
            if fallback:
                font, fallback_size = fallback
                font_cell.value = font
                if not size_val:
                    size_cell.value = fallback_size

        if font not in allowed_fonts:
            if font:
                font_cell.fill = orange_fill
            size_cell.fill = red_fill
        else:
            try:
                size = float(size_cell.value)
                min_size, max_size = allowed_fonts[font]
                if not (min_size <= size <= max_size):
                    size_cell.fill = red_fill
            except (TypeError, ValueError):
                size_cell.fill = red_fill

        font_color = str(font_color_cell.value or "").strip().upper()
        if font_color and font_color not in allowed_font_colors:
            font_color_cell.fill = yellow_fill

    wb.save(excel_path)
    return "Validation updated successfully."

def color_slide_point_comments(excel_path):
    comment_color_map = {
        "Perfect match (copied)": "ff0000",
        "Chunked properly": "87E179",
        "No strong match": "FF9999",
        "Partially matching": "9ADFE6",
        "No VO content": "8B0000"
    }
    wb = load_workbook(excel_path)
    if "Slide Point Analysis" not in wb.sheetnames:
        return

    ws = wb["Slide Point Analysis"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "Comment" not in header:
        return

    comment_col_idx = header.index("Comment") + 1
    for row in ws.iter_rows(min_row=2, min_col=comment_col_idx, max_col=comment_col_idx):
        cell = row[0]
        comment = str(cell.value).strip() if cell.value else ""
        if comment in comment_color_map:
            cell.fill = PatternFill(start_color=comment_color_map[comment], end_color=comment_color_map[comment], fill_type="solid")

    wb.save(excel_path)

def highlight_animations(excel_path):
    wb = load_workbook(excel_path)
    if "Animation QC" not in wb.sheetnames:
        return

    ws = wb["Animation QC"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "Animation Type" not in header:
        return

    anim_col_idx = header.index("Animation Type") + 1
    for row in ws.iter_rows(min_row=2):
        cell = row[anim_col_idx - 1]
        value = str(cell.value).lower().strip() if cell.value else ""
        if "fade" in value or "wipe" in value:
            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        elif value == "unknown":
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    wb.save(excel_path)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_files():
    file_a = request.files['file_a']
    file_b = request.files['file_b']
    filename_a = secure_filename(file_a.filename)
    filename_b = secure_filename(file_b.filename)
    path_a = os.path.join(UPLOAD_FOLDER, filename_a)
    path_b = os.path.join(UPLOAD_FOLDER, filename_b)
    file_a.save(path_a)
    file_b.save(path_b)

    ungrouped_path_b = os.path.join(UPLOAD_FOLDER, "ungrouped_" + filename_b)
    ungroup_shapes_in_ppt(path_b, ungrouped_path_b)

    output_filename = f"{os.path.splitext(filename_b)[0]}_QC_Report.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    df_animation = run_animation_qc(ungrouped_path_b)
    df_slide_point, df_summary = run_chunking_qc_with_animation(ungrouped_path_b)
    df_notes_a, df_notes_b, df_cmp, df_qc = run_notes_validation(path_a, ungrouped_path_b)
    df_text_rules = run_text_rules_validation(df_qc)

    df_slide_point = clean_illegal_excel_chars(df_slide_point)
    df_summary = clean_illegal_excel_chars(df_summary)
    df_animation = clean_illegal_excel_chars(df_animation)
    df_notes_a = clean_illegal_excel_chars(df_notes_a)
    df_notes_b = clean_illegal_excel_chars(df_notes_b)
    df_cmp = clean_illegal_excel_chars(df_cmp)
    df_qc = clean_illegal_excel_chars(df_qc)
    df_text_rules = clean_illegal_excel_chars(df_text_rules)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_slide_point.to_excel(writer, sheet_name="Slide Point Analysis", index=False)
        df_summary.to_excel(writer, sheet_name="Summary Review", index=False)
        df_animation.to_excel(writer, sheet_name="Animation QC", index=False)
        df_notes_a.to_excel(writer, sheet_name="File A Notes", index=False)
        df_notes_b.to_excel(writer, sheet_name="File B Notes", index=False)
        df_cmp.to_excel(writer, sheet_name="Comparison Results", index=False)
        df_qc.to_excel(writer, sheet_name="Quality Check", index=False)
        df_text_rules.to_excel(writer, sheet_name="Text Rules Check", index=False)

    color_slide_point_comments(output_path)
    highlight_animations(output_path)
    update_font_validation_with_fallback(output_path)

    #  Add new sheet summarizing all QC issues
    generate_qc_summary(output_path)

    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
