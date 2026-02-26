import pandas as pd
from openpyxl import load_workbook

# Approved color sets
APPROVED_FONT_COLORS = {"#000000", "#FFFFFF", "#F26722"}
APPROVED_FILL_COLORS = {
    "#F26722", "#0045C0", "#FDD900", "#B9CB00",
    "#3B4096", "#CCC1FF", "#27BDBB", "#117673"
}

def generate_qc_summary(excel_path):
    wb = load_workbook(excel_path)
    all_issues = []

    def is_valid(val):
        return pd.notna(val) and str(val).strip().lower() not in ["", "nan"]

    # --- 1. Slide Point Analysis ---
    if "Slide Point Analysis" in wb.sheetnames:
        df = pd.read_excel(excel_path, sheet_name="Slide Point Analysis")
        for _, row in df.iterrows():
            comment = row.get("Comment")
            slide_point = row.get("Slide Point")
            if is_valid(slide_point):
                if comment == "No VO content":
                    all_issues.append({
                        "Slide Number": row.get("Slide Number"),
                        "Issue Type": "No VO Content",  # Specific issue for No VO content
                        "Description": slide_point,
                        "Shape ID": ""
                    })
                elif comment == "Partially matching":
                    all_issues.append({
                        "Slide Number": row.get("Slide Number"),
                        "Issue Type": "Partially Matching",  # Specific issue for Partially matching
                        "Description": slide_point,
                        "Shape ID": ""
                    })
                elif comment == "No strong match":
                    all_issues.append({
                        "Slide Number": row.get("Slide Number"),
                        "Issue Type": "Chunking",
                        "Description": slide_point,
                        "Shape ID": ""
                    })

    # --- 2. Animation QC ---
    if "Animation QC" in wb.sheetnames:
        df = pd.read_excel(excel_path, sheet_name="Animation QC")
        for _, row in df.iterrows():
            animation_type = str(row.get("Animation Type", "")).strip().lower()
            slide = row.get("Slide Number", "")
            shape = row.get("Shape Name / Table Cell", "")
            text = row.get("Text", "")

            if animation_type == "unknown":
                all_issues.append({
                    "Slide Number": slide,
                    "Issue Type": "Unknown Animation",
                    "Description": text if is_valid(text) else "(No text)",
                    "Shape ID": shape
                })

    # --- 3. Text Rules Check ---
    if "Text Rules Check" in wb.sheetnames:
        df = pd.read_excel(excel_path, sheet_name="Text Rules Check", dtype=str)
        for _, row in df.iterrows():
            text = row.get("Extracted Text", "")
            slide = row.get("Slide Number", "")
            shape = row.get("Shape Name / Table Cell", "")

            # Contraction check
            if row.get("Contraction Used") == "Yes" and is_valid(text):
                all_issues.append({"Slide Number": slide, "Issue Type": "Contraction", "Description": text, "Shape ID": shape})

            # US English spelling check
            if is_valid(row.get("US English Used")):
                all_issues.append({"Slide Number": slide, "Issue Type": "US English", "Description": row.get("US English Used"), "Shape ID": shape})

            # Extra space check
            if row.get("Extra Space") == "Yes" and is_valid(text):
                all_issues.append({"Slide Number": slide, "Issue Type": "Extra Space", "Description": text, "Shape ID": shape})

            # Ending period check
            if row.get("Ending Period Used") == "Yes" and is_valid(text):
                all_issues.append({"Slide Number": slide, "Issue Type": "Ending Period", "Description": text, "Shape ID": shape})

            # Mid-sentence period check
            if row.get("Mid Sentence Period Used") == "Yes" and is_valid(text):
                all_issues.append({"Slide Number": slide, "Issue Type": "Mid Sentence Period", "Description": text, "Shape ID": shape})

    # --- 4. Quality Check ---
    if "Quality Check" in wb.sheetnames:
        df = pd.read_excel(excel_path, sheet_name="Quality Check")
        for _, row in df.iterrows():
            slide = row.get("Slide Number", "")
            text = row.get("Extracted Text", "")
            shape = row.get("Shape Name / Table Cell", "")

            font_name = str(row.get("Font Name", "")).strip()
            font_color = str(row.get("Font Color Hex", "")).upper().strip()
            fill_color = str(row.get("Fill Color Hex", "")).upper().strip()

            # Font check: Unapproved font
            if is_valid(font_name) and font_name not in [
                "HelveticaNowDisplay Medium", "Queens Medium", "HelveticaNowDisplay Black", "Consolas", "Cambria math"
            ]:
                all_issues.append({
                    "Slide Number": slide,
                    "Issue Type": "Unapproved Font",
                    "Description": font_name,
                    "Shape ID": shape
                })

            # Font size check (ranges)
            try:
                size = float(row.get("Font Size", 0))
                if font_name == "Queens Medium" and size != 35:
                    all_issues.append({
                        "Slide Number": slide,
                        "Issue Type": "Font Size Mismatch",
                        "Description": f"{font_name} : {size}",
                        "Shape ID": shape
                    })
                elif font_name == "HelveticaNowDisplay Medium" and not (24 <= size <= 27):
                    all_issues.append({
                        "Slide Number": slide,
                        "Issue Type": "Font Size Mismatch",
                        "Description": f"{font_name} : {size}",
                        "Shape ID": shape
                    })
                elif font_name == "HelveticaNowDisplay Black" and not (70 <= size <= 92.5):
                    all_issues.append({
                        "Slide Number": slide,
                        "Issue Type": "Font Size Mismatch",
                        "Description": f"{font_name} : {size}",
                        "Shape ID": shape
                    })
            except:
                pass

            # Font color mismatch
            if is_valid(font_color) and font_color not in APPROVED_FONT_COLORS:
                all_issues.append({
                    "Slide Number": slide,
                    "Issue Type": "Unapproved Font Color",
                    "Description": font_color,
                    "Shape ID": shape
                })

            # Fill color mismatch
            if is_valid(fill_color) and fill_color not in APPROVED_FILL_COLORS:
                all_issues.append({
                    "Slide Number": slide,
                    "Issue Type": "Unapproved Fill Color",
                    "Description": fill_color,
                    "Shape ID": shape
                })

    # --- Final Cleanup ---
    df_final = pd.DataFrame(all_issues)
    df_final = df_final[df_final["Description"].apply(lambda x: is_valid(x))]
    df_final.drop_duplicates(inplace=True)
    df_final["Slide Number"] = pd.to_numeric(df_final["Slide Number"], errors="coerce")
    df_final.dropna(subset=["Slide Number"], inplace=True)
    df_final["Slide Number"] = df_final["Slide Number"].astype(int)
    df_final.sort_values(by=["Slide Number"], inplace=True)

    # --- Write QC Points Sheet ---
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_final.to_excel(writer, sheet_name="QC Points", index=False)

    print("QC Points sheet generated successfully.")
